import cv2
import os
import face_recognition
from datetime import datetime
from openpyxl import Workbook, load_workbook

class F_reconigtion:

   def __init__(self):
        self.registrados = set()

   def ositopanda(self, numero_cuenta):  # Agregar numero_cuenta como parámetro
        # Codificar los rostros
        imageFacesPath = "C:\\Users\erick\Desktop\VELOZER\PROYECTO FINAL\Asistencia_inteligente\\faces"
        facesEncodings = []
        facesNames = []
        for file_name in os.listdir(imageFacesPath):
            image = cv2.imread(imageFacesPath + "/" + file_name)
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            f_coding = face_recognition.face_encodings(image, known_face_locations=[(0, 150, 150, 0)])[0]
            facesEncodings.append(f_coding)
            facesNames.append(file_name.split(".")[0])

        # Crear archivo Excel si no existe
        if not os.path.exists('Asistencia.xlsx'):
            wb = Workbook()
            ws = wb.active
            ws.append(["Nombre y Numero de Cuenta", "Numero de Cuenta", "Fecha y Hora"])
            wb.save('Asistencia.xlsx')

        # Cargar archivo Excel
        wb = load_workbook('Asistencia.xlsx')
        ws = wb.active

        # VIDEO
        cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        # Detector facial
        facecCassif = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

        while True:
            ret, frame = cap.read()
            if ret == False:
                break
            frame = cv2.flip(frame, 1)
            orig = frame.copy()
            faces = facecCassif.detectMultiScale(frame, 1.1, 5)
            for (x, y, w, h) in faces:
                face = orig[y:y + h, x:x + w]
                face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
                actual_face_encoding = face_recognition.face_encodings(face, known_face_locations=[(0, w, h, 0)])[0]
                result = face_recognition.compare_faces(facesEncodings, actual_face_encoding)
                print(result)
                if True in result:
                    index = result.index(True)
                    name = facesNames[index]
                    color = (125, 220, 0)
                    if name not in self.registrados:
                        # Guardar datos en Excel solo si la persona no ha sido registrada antes
                        now = datetime.now()
                        date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
                        ws.append([name, numero_cuenta, date_time])  # Agregar numero_cuenta al registro
                        self.registrados.add(name)
                else:
                    name = "Desconocido"
                    color = (50, 50, 225)

                cv2.rectangle(frame, (x, y + h), (x + w, y + h + 30), color, -1)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, name, (x, y + h + 25), 2, 1, (255, 255, 255), 2, cv2.LINE_AA)

            cv2.imshow("Frame", frame)
            k = cv2.waitKey(1) & 0xFF
            if k == 27:
                break

        # Guardar y cerrar Excel
        wb.save('Asistencia.xlsx')
        wb.close()

        cap.release()
        cv2.destroyAllWindows()